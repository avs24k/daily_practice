"""webdrivwer invoke"""
import openpyxl
from selenium import webdriver
from selenium.webdriver import ActionChains, Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait
# service_obj = Service()
# driver = webdriver.Chrome(service=service_obj)
# driver.get("https://www.amazon.in/?&ext_vrnc=hi&tag=googhydrabk1-21&ref=pd_sl_7hz2t19t5c_e&adgrpid=58355126069&hvpone"
#            "=&hvptwo=&hvadid=678711876615&hvpos=&hvnetw=g&hvrand=259229048841897065&hvqmt=e&hvdev=c&hvdvcmdl=&hvlocint"
#            "=&hvlocphy=9181907&hvtargid=kwd-10573980&hydadcr=14453_2371562")



"""locators"""
# Name
# driver.find_element(By.NAME,"site-search").click()

# tag_name/Anchor_tag_name/tag_ID
# driver.find_element(By.ID,"nav-search-dropdown-card").click()

# Class_Name
# driver.find_element(By.CLASS_NAME,"nav-left").click()

# CSS_Selector
# driver.find_element(By.CSS_SELECTOR,"label[id='searchDropdownDescription']").click()

# Xpath
# driver.find_element(By.XPATH,"//label[@id='searchDropdownDescription']").click()

# Link_Text
# driver.find_element(By.LINK_TEXT,"Skip to main content").click()

# partial_Link_Text
# driver.find_element(By.PARTIAL_LINK_TEXT,"Skip to main").click()

# Dynamic Xpath
# driver.find_element(By.XPATH,"//span[contains(text(),'iPhone 13 (128GB)')]").click()

"""Handling Download"""

# options = webdriver.ChromeOptions()
# prefs = {'download.default_directory':'Downloads'#download_folder_path}
# options.add_experimental_option('prefs',prefs)


"""Waits"""

# driver.implicitly_wait(3)
# wait = WebDriverWait(driver, 10)
# wait.until(expected_conditions._element_if_visible(driver.find_element(By.XPATH,"//form[@id='nav-search-bar-form']")).send_keys("avinash")

"""Cookies"""

# driver.implicitly_wait(10)
# driver.get("https://rahulshettyacademy.com/AutomationPractice/")
# driver.get("about:blank")

# delt = driver.delete_cookie("bypass_utm")
# cooki = driver.get_cookies()
# driver.add_cookie({'name':'bypass_utm'})

"""Static DropDown"""

# dropdown = Select(driver.find_element(By.XPATH,"//label[@id='searchDropdownDescription']"))
# dropdown.select_by_value()
# dropdown.select_by_index()
# dropdown.select_by_visible_text()


"""Java Script Alerts Popups"""
# driver.find_element(By.XPATH,"//label[@id='searchDropdownDescription']").click()
# alert = driver.switch_to.alert
# alert.accept()
# alert.dismiss()


"""Mouse Interactions"""
# action = ActionChains(driver)
# action.click()

'''How to send ENTER/TAB keys in WebDriver?'''
    # action.send_keys(Keys.ENTER)

# context_click means right click
    # action.context_click(driver.find_element(By.XPATH,"//label[@id='searchDropdownDescription']").perform()

# drag and drop operation
    # Find the draggable element
    # draggable = driver.find_element_by_id("draggable")

    # Find the droppable element
    # droppable = driver.find_element_by_id("droppable")

    # Create an ActionChains object
    # actions = ActionChains(driver)

    # Perform drag-and-drop
    # actions.drag_and_drop(draggable, droppable).perform()


# action.double_click(driver.find_element(By.XPATH,"//label[@id='searchDropdownDescription']"))
# action.scroll_to_element(driver.find_element(By.XPATH,"//label[@id='searchDropdownDescription']"))
# action.click_and_hold(driver.find_element(By.XPATH,"//label[@id='searchDropdownDescription']"))


"""Multiple Window Handling"""
# windows_handler = driver.window_handles
# driver.switch_to.window(windows_handler[1])

"""Frames Handling"""

# driver.switch_to.frame(driver.find_elements())
# driver.switch_to.default_content()

"""driver navigation"""
# driver.back()
# driver.forward()
# driver.refresh()


"""Scroll Page"""
# time.sleep(3)
# driver.execute_script("window.scrollTo(0,500)")
# driver.execute_script("window.scrollTo(0,document.body.scrollHeight)")
# X = driver.find_element(By.XPATH,"//legend[normalize-space()='Mouse Hover Example']")
# driver.execute_script("arguments[0].scrollIntoView();",X)

"""java script executor"""

# driver.get_screenshot_as_file('avi.png')


"""Read and Write file, file Handling, context collector"""

# if we are opening file using only "open Statement" then we need to close it manually"

# file0 = open("file_zoath")
# print(file0.readlines())
# file0.close()


# if we are using "with statement" then we don't need to worry about to close the file. it is automatically close
# when operation is done

# with open ("file_path","r") as file:
#     read_file = file.readlines()
#     print(read_file)
#
# with open("file_path","w") as file1:
#     write_file = file1.write("avinash"

"""Raise Exception"""

# def myfun(n):
#     try:
#         result = 10 / n
#         return result
#     except ZeroDivisionError as e:
#         return e
#
#
# print(myfun(0))

"""Write the code for Reading and Writing to Excel through Selenium"""

# workbooks = openpyxl.load_workbook('C:\\Users\Vikky\Downloads\LoginData (1).xlsx')
# sheet1 = workbooks.active
#
# for i in sheet1.iter_rows(min_row=1, max_row=sheet1.max_row, min_col=1, max_col=sheet1.max_column):
#     for j in i:
#         print(j.value,end='\t')
#     print()